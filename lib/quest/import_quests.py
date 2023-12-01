import io
from contextlib import redirect_stdout
from typing import Dict, Optional, TypedDict

from firebase_admin import firestore
from google.cloud.firestore import Client as FirestoreClient
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class Quest(TypedDict):
    journey_id: int
    chapter_id: int
    quest_id: int
    data: Dict


def parse_quest_row(
    worksheet: Worksheet, row: int, column_names: Dict[str, int]
) -> Optional[Quest]:
    journey_id = int(worksheet.cell(row, column_names["EveryQuestJourney"]).value)  # type: ignore
    chapter_id = int(worksheet.cell(row, column_names["EveryQuestChapter"]).value)  # type: ignore
    quest_id = int(worksheet.cell(row, column_names["EveryQuestId"]).value)  # type: ignore

    quest_str = f"{journey_id}:{chapter_id}:{quest_id}"
    print(f"importing quest data for {quest_str}\n")

    choices = []
    data = {
        "after": "",
        "before": "",
        "chapterID": str(chapter_id),
        "contentType": worksheet.cell(row, column_names["EveryQuestContentType"]).value,
        "journeyID": str(journey_id),
        "title": worksheet.cell(row, column_names["EveryQuestTitle"]).value,
        "tooltipEmoji": worksheet.cell(
            row, column_names["TooltipManagementTooltipEmoji"]
        ).value,
        "successTooltip": worksheet.cell(
            row, column_names["TooltipManagementTooltipText"]
        ).value,
    }

    quest_type = worksheet.cell(row, column_names["EveryQuestContentType"]).value

    match quest_type:
        case "video":
            url = worksheet.cell(row, column_names["VideoQuestVideoUrl"]).value
            if url is None:
                print(
                    f"warning: video quest type requires a video URL, skipping {quest_str}\n"
                )
            data["video"] = {
                "duration": 0,
                "thumbnailURL": worksheet.cell(
                    row, column_names["VideoQuestThumbnailUrl"]
                ).value,
                "url": url,
            }
        case "riff":
            question = worksheet.cell(row, column_names["RiffQuestion"]).value
            if question is None:
                print(
                    f"warning: riff quest type requires a question, skipping {quest_str}\n"
                )
                return None
            data["riff"] = {
                "question": question,
                "tip": worksheet.cell(row, column_names["RiffTip"]).value,
            }
        case "multiChoiceQuestion":
            for index in range(0, 6):
                if worksheet.cell(
                    row, index + column_names["MultiChoiceAndMultiChoiceAnswerChoice1"]
                ).value:
                    option = worksheet.cell(
                        row,
                        index + column_names["MultiChoiceAndMultiChoiceAnswerChoice1"],
                    ).value
                    choice = {
                        "id": index,
                        "option": option,
                    }
                    choices.append(choice)

            data["answerQuestID"] = worksheet.cell(
                row, column_names["AnswerQuestID"]
            ).value
            data["question"] = {"choices": choices}

            correct_answer = str(
                worksheet.cell(
                    row, column_names["MultiChoiceAndMultiChoiceAnswerCorrectAnswer"]
                ).value
            )
            if not correct_answer:
                if "," not in correct_answer:
                    correct_answer = [int(correct_answer) - 1]
                else:
                    correct_answer = correct_answer.split(",")
                    # subtract 1 from the correct answer elements for zero based indexing
                    correct_answer = [int(x) - 1 for x in correct_answer]
                data["correctAnswer"] = correct_answer
        case "multiChoiceAnswer":
            for index in range(0, 6):
                if worksheet.cell(
                    row, index + column_names["MultiChoiceAndMultiChoiceAnswerChoice1"]
                ).value:
                    choice = {
                        "id": index,
                        "option": worksheet.cell(
                            row,
                            index
                            + column_names["MultiChoiceAndMultiChoiceAnswerChoice1"],
                        ).value,
                        "totalAnswer": 0,
                    }
                    choices.append(choice)

            data["total"] = 0
            data["question"] = {"choices": choices}
        case "message":
            data["imageUrl"] = worksheet.cell(
                row, column_names["MessageQuestImageurl"]
            ).value
            data["body"] = worksheet.cell(row, column_names["MessageQuestBody"]).value
        case "communityResponse":
            data["title"] = worksheet.cell(
                row, column_names["QuestionCommunityResponse"]
            ).value
        case "madlib":
            data["question"] = worksheet.cell(row, column_names["MadLibQuestion"]).value
        case "tapAndDrag":
            choicesList = str(
                worksheet.cell(row, column_names["Tap&DragAnswers"]).value
            ).split(",")
            choices = []
            for index in range(0, len(choicesList)):
                choice = {"id": index, "text": choicesList[index]}
                choices.append(choice)

            zones = []
            for index in range(0, 3):
                if worksheet.cell(row, index + column_names["Tap&DragId1"]).value:
                    zone = {
                        "id": index,
                        "text": worksheet.cell(
                            row, index + column_names["Tap&DragId1"]
                        ).value,
                    }
                    zones.append(zone)

            data["answers"] = choices
            data["zones"] = zones
        case _:
            print(
                f"warning: invalid quest type: found '{quest_type}', skipping {quest_str}\n"
            )
            return None

    return {
        "journey_id": journey_id,
        "chapter_id": chapter_id,
        "quest_id": quest_id,
        "data": data,
    }


def import_quest_data(db: FirestoreClient, worksheet: Worksheet) -> None:
    column_names = {}
    current = 1
    columns = worksheet.iter_cols(
        min_col=0, min_row=3, max_col=worksheet.max_column, max_row=3
    )
    for column in columns:
        column_names[column[0].value] = current
        current += 1

    for row in range(4, worksheet.max_row):
        if worksheet.cell(row, 1).value is None:
            break

        quest = None
        try:
            quest = parse_quest_row(worksheet, row, column_names)
        except ValueError as e:
            print(f"{e}\n")

        if quest is None:
            continue

        journey_id = quest["journey_id"]
        chapter_id = quest["chapter_id"]
        quest_id = quest["quest_id"]
        data = quest["data"]

        path = f"journeys/{journey_id}/chapters/{chapter_id}/quests"
        db.collection(path).document(str(quest_id)).set(data)


def import_chapter_metadata(db: FirestoreClient, worksheet: Worksheet) -> None:
    column_names = {}
    current = 1
    for column in worksheet.iter_cols(
        min_col=0, min_row=1, max_col=worksheet.max_column, max_row=1
    ):
        column_names[column[0].value] = current
        current += 1

    for row in range(2, worksheet.max_row):
        journey_id = worksheet.cell(row, column_names["journey"]).value
        chapter_id = worksheet.cell(row, column_names["chapter"]).value

        if journey_id is None:
            break

        print(f"importing chapter metadata for {journey_id}:{chapter_id}\n")

        data = {
            "name": worksheet.cell(row, column_names["name"]).value,
            "chapterNumber": worksheet.cell(row, column_names["chapterNumber"]).value,
            "author": worksheet.cell(row, column_names["author"]).value,
            "title": worksheet.cell(row, column_names["Title"]).value,
            "subTitle": worksheet.cell(row, column_names["subTitle"]).value,
            "description": worksheet.cell(row, column_names["description"]).value,
            "image": worksheet.cell(row, column_names["image"]).value,
        }

        path = f"journeys/{journey_id}/chapters"
        db.collection(path).document(str(chapter_id)).set(data)


def import_journey_metadata(db: FirestoreClient, worksheet: Worksheet) -> None:
    column_names = {}
    current = 1
    for column in worksheet.iter_cols(
        min_col=0, min_row=1, max_col=worksheet.max_column, max_row=1
    ):
        column_names[column[0].value] = current
        current += 1

    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row, 1).value is None:
            break

        try:
            journey_id = int(worksheet.cell(row, column_names["journey"]).value)  # type: ignore
        except ValueError as e:
            print(f"{e}\n")
            continue

        print(f"importing journey metadata for {journey_id}\n")

        data = {
            "name": worksheet.cell(row, column_names["name"]).value,
            "image": worksheet.cell(row, column_names["image"]).value,
            "author": worksheet.cell(row, column_names["author"]).value,
        }

        db.collection("journeys").document(str(journey_id)).set(data)


def import_quests(import_file):
    db = firestore.client()
    with io.StringIO() as buffer, redirect_stdout(buffer):
        worksheet_filename = import_file
        workbook = load_workbook(worksheet_filename)

        quest_worksheet_name = "quest_data"
        chapter_metadata_worksheet_name = "chapter_metadata"
        journey_metadata_worksheet_name = "journey_metadata"

        quest_worksheet = workbook[quest_worksheet_name]
        chapter_metadata_worksheet = workbook[chapter_metadata_worksheet_name]
        journey_metadata_worksheet = workbook[journey_metadata_worksheet_name]

        print("Importing quests...\n")
        import_quest_data(db, quest_worksheet)

        print("\nImporting chapter metadata...\n")
        import_chapter_metadata(db, chapter_metadata_worksheet)

        print("\nImporting journey metadata...\n")
        import_journey_metadata(db, journey_metadata_worksheet)

        print("\nFinished import!")
        output = buffer.getvalue()
    return output
